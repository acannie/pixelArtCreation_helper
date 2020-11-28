from PIL import Image
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import math
import pandas as pd
from sklearn.cluster import KMeans
import colorsys
import sys

# 正方形にトリミング（中央寄せ）
def trimmingImg(im=[[[0,0,0]]]):
    IMG_HEIGHT = im.shape[0]
    IMG_WIDTH = im.shape[1]

    LONG_SIDE_AXIS = None
    SHORT_SIDE_AXIS = None

    if IMG_HEIGHT < IMG_WIDTH: # 画像が横長
        LONG_SIDE_AXIS = 1
        SHORT_SIDE_AXIS = 0
    else: # 画像が縦長
        LONG_SIDE_AXIS = 0
        SHORT_SIDE_AXIS = 1
    
    LONG_SIDE_LENGTH = max(IMG_HEIGHT, IMG_WIDTH)
    SHORT_SIDE_LENGTH = min(IMG_HEIGHT, IMG_WIDTH)

    TARGET_LENGTH = math.floor(SHORT_SIDE_LENGTH / ART_WH) # 1辺の長さが ART_WIDTH の倍数になるよう切り落とし
    SHARD = LONG_SIDE_LENGTH - TARGET_LENGTH

    im = np.delete(im, np.s_[math.floor((SHORT_SIDE_LENGTH - TARGET_LENGTH) / 2):], SHORT_SIDE_AXIS)
    im = np.delete(im, np.s_[:math.ceil(SHARD / 2)], SHORT_SIDE_AXIS)
    im = np.delete(im, np.s_[LONG_SIDE_LENGTH -
                            math.floor((LONG_SIDE_LENGTH - TARGET_LENGTH) / 2):], LONG_SIDE_AXIS)
    im = np.delete(im, np.s_[:math.ceil((LONG_SIDE_LENGTH - TARGET_LENGTH) / 2)], LONG_SIDE_AXIS)

# RGBの数値が1桁の場合冒頭に0をつけて2桁にして返す
def addZero(strHexNum='0xAA'):
    if len(strHexNum) == 3:
        return '0x0' + strHexNum[2]
    elif len(strHexNum) == 4:
        return strHexNum
    else:
        print('error!')
        sys.exit()

# RGBをカラーコードに変換
def RGBtoColorCode(RGB=[0, 0, 0]):
    color_code = ''
    for color in range(COLOR_VARIETY):
        color_code += str(addZero(hex(RGB[color])))
    color_code = color_code.replace('0x', '') # '0xrr0xgg0xbb' を 'rrggbb' にする
    return color_code


# 明度 (0 < V < 255)
def getBrightness(RGB=[0,0,0]):
    max_color = max(RGB)
    V = max_color
    return V

# 背景色によって文字色を変える
def getFontColor(RGB=[0,0,0]):
    if getBrightness(RGB) > 200:
        return '000000'
    else:
        return 'FFFFFF'

# デザインゾーンの枠線を追加
def drawRuledLine(ws=openpyxl.Workbook().worksheets[0]):
    thick = Side(style='thick', color='000000')
    for i in range(3, ART_WH + 1):
        ws.cell(row=2, column=i).border = Border(top=thick)
    for i in range(3, ART_WH + 1):
        ws.cell(row=i, column=2).border = Border(left=thick)
    for i in range(3, ART_WH + 1):
        ws.cell(row=i, column=ART_WH+1).border = Border(right=thick)
    for i in range(3, ART_WH + 1):
        ws.cell(row=ART_WH+1, column=i).border = Border(bottom=thick)
    ws.cell(row=2, column=2).border = Border(top=thick, left=thick)
    ws.cell(row=2, column=ART_WH+1).border = Border(top=thick, right=thick)
    ws.cell(row=ART_WH+1, column=2).border = Border(left=thick, bottom=thick)
    ws.cell(row=ART_WH+1, column=ART_WH +
            1).border = Border(right=thick, bottom=thick)

    # デザインゾーンの破線を追加
    mediumDashed = Side(style='mediumDashed', color='000000')
    for i in range(3, ART_WH + 1):
        ws.cell(row=round(ART_WH / 2) + 1, column=i).border = Border(bottom=mediumDashed)
    for i in range(3, ART_WH + 1):
        ws.cell(row=i, column=round(ART_WH / 2) + 1).border = Border(right=mediumDashed)
    ws.cell(row=round(ART_WH / 2) + 1, column=round(ART_WH / 2) + 1).border = Border(right=mediumDashed, bottom=mediumDashed)
    ws.cell(row=round(ART_WH / 2) + 2, column=round(ART_WH / 2) + 2).border = Border(top=mediumDashed, left=mediumDashed)
    ws.cell(row=2, column=round(ART_WH / 2) + 1).border = Border(top=side1, right=mediumDashed)
    ws.cell(row=round(ART_WH / 2) + 1, column=2).border = Border(left=side1, bottom=mediumDashed)
    ws.cell(row=round(ART_WH / 2) + 1, column=ART_WH + 1).border = Border(right=side1, bottom=mediumDashed)
    ws.cell(row=ART_WH + 1, column=round(ART_WH / 2) + 1).border = Border(bottom=side1, right=mediumDashed)


# あつ森の仕様
ART_WH = 32
ART_PALETTE = 15
ART_COLOR_VARIETY = 30
ART_BRIGHTNESS_VARIERY = 15
ART_VIVIDNESS_VARIETY = 15

# ファイルのインポート
if len(sys.argv) != 2:
    sys.exit()

input_file = sys.argv[1]
im = np.array(Image.open('figure/' + input_file))

trimmingImg(im)

# 切り取った画像を保存
Image.fromarray(im.astype(np.uint8)).save('squaredFigure/' + input_file)

# 画像のサイズを取得
IMG_HEIGHT = im.shape[0]
IMG_WIDTH = im.shape[1]
COLOR_VARIETY = im.shape[2]

# ワークブックを作成
wb = openpyxl.Workbook()
ws = wb.worksheets[0]

drawRuledLine()


# デザインゾーンのセルを正方形に整形
for x in range(1, ART_WH+1):
    ws.column_dimensions[get_column_letter(x+1)].width = 4
for y in range(1, ART_WH+1):
    ws.row_dimensions[y+1].height = 22

# art_wh * art_wh のドット絵を生成
RGB_art_wh = []
df_all_color = pd.DataFrame(columns=['R', 'G', 'B'])

color_count = 0

for r in range(ART_WH):
    RGB_art_wh.append([])

    for c in range(ART_WH):
        RGB_ave = [0, 0, 0]

        # n×nピクセルの平均RGBをセルの背景色とする
        n = int(IMG_WIDTH / ART_WH)
        for i in range(n):
            for j in range(n):
                for color in range(im.shape[2]):
                    RGB_ave[color] += im[n * r + i][n * c + j][color]

        for color in range(im.shape[2]):
            RGB_ave[color] = round(RGB_ave[color] / n**2)

        RGB_art_wh[r].append(RGB_ave)
        df_all_color = df_all_color.append(
            {'R': RGB_ave[0], 'G': RGB_ave[1], 'B': RGB_ave[2]}, ignore_index=True)
        color_count += 1

cust_array = np.array([df_all_color['R'].tolist(),
                       df_all_color['G'].tolist(),
                       df_all_color['B'].tolist(), ], np.int32)

# 行列を転置
cust_array = cust_array.T

# クラスタ分析(クラスタ数=15)で色の種類を絞る
pred = KMeans(n_clusters=15).fit_predict(cust_array)
df_all_color['color'] = pred

cluster_RGB_list = []
HSV_list = []
for i in range(15):
    cluster_RGB = df_all_color[df_all_color['color'] == i].mean()

    cluster_RGB_list.append([])
    cluster_RGB_list[i].append(round(cluster_RGB[0]))
    cluster_RGB_list[i].append(round(cluster_RGB[1]))
    cluster_RGB_list[i].append(round(cluster_RGB[2]))
    
    # HSV_list[i].append(round(getHue(list(cluster_RGB))/360 * 30))
    # HSV_list[i].append(round(getSaturation(list(cluster_RGB))*15))
    # HSV_list[i].append(round(getBrightness(list(cluster_RGB))/255*15))

    hsv = colorsys.rgb_to_hsv(cluster_RGB[0], cluster_RGB[1], cluster_RGB[2])
    HSV_list.append([])
    HSV_list[i].append(math.ceil(hsv[0] * 30))
    HSV_list[i].append(math.ceil(hsv[1] * 15))
    HSV_list[i].append(math.ceil(hsv[2]/255 * 15))
    # HSV_list[i].append(round(getSaturation(list(cluster_RGB))*15))
    # HSV_list[i].append(round(getBrightness(list(cluster_RGB))/255*15))


# ゲージ
GAUGE_WIDTH = 30
now_pacentage = 0 # 範囲: 0 - 100

# 描画
for index, row in df_all_color.iterrows():
    color_code = RGBtoColorCode(cluster_RGB_list[row['color']])
    fill = PatternFill(patternType='solid', fgColor=color_code)
    row_num = math.ceil((index + 1) / ART_WH) + 1
    col_num = index % ART_WH + 2
    ws.cell(row=row_num, column=col_num).fill = fill
    ws.cell(row=row_num, column=col_num).value = row['color'] + 1
    ws.cell(row=row_num, column=col_num).font = openpyxl.styles.fonts.Font(color=getFontColor(cluster_RGB_list[row['color']]))


# カラーパレット
ws['AJ2'].value = '色相'
ws['AK2'].value = '彩度'
ws['AL2'].value = '明度'

for i in range(ART_PALETTE):
    color_code = RGBtoColorCode(cluster_RGB_list[i])
    fill = PatternFill(patternType='solid', fgColor=color_code)
    ws.cell(row=i+3, column=ART_WH + 3).fill = fill
    ws.cell(row=i+3, column=ART_WH + 3).value = i + 1
    ws.cell(row=i+3, column=ART_WH + 3).font = openpyxl.styles.fonts.Font(color=getFontColor(cluster_RGB_list[i]))

    ws.cell(row=i+3, column=ART_WH + 4).value = HSV_list[i][0]
    ws.cell(row=i+3, column=ART_WH + 5).value = HSV_list[i][1]
    ws.cell(row=i+3, column=ART_WH + 6).value = HSV_list[i][2]

# ファイルのエクスポート
print('')
print("saving...")

output_file = input_file.replace('.jpg', '.xlsx')
wb.save('result/' + output_file)
wb.close()

print("complete!")
